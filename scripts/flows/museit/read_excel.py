from prefect import flow, task
from openpyxl.reader.excel import load_workbook
import boto3
import requests
import json
from prefect.logging import get_run_logger
import random
import base64


@task
def map_line_to_metadata(row, mapping, template):
    logger = get_run_logger()
    asset = {}
    asset['title'] = row[0]
    asset['author'] = row[1]
    asset['description'] = row[2] if row[2] else "Placeholder"
    print("Row: ", row)
    print("Asset: ", asset)
    print("Mapping: ", mapping)
    print("Template: ", template)
    response = requests.post(
        url='http://dataversemapper:8099/mapper/',
        json={
            'metadata': asset,
            'template': template,
            'mapping': mapping,
        }
    )
    logger.info("Response: ", response.json())
    return response

@task
def retrieve_file_for_metadata(row):
    access_key = 'zGnGNFec3DKTXiN790kZ'
    bucketname = 'museit'
    secret_key = 'eBZts8xTc3wbU1UEe5E0fHufTiZtBqwMItFbC9oC'
    filename = row[3]
    minio_client = boto3.client('s3', endpoint_url='http://nginxminio:9000', aws_access_key_id=access_key, aws_secret_access_key=secret_key)
    filedata = minio_client.get_object(Bucket=bucketname, Key=filename)
    return filedata['Body'].read()
    #the bagpipe is angry

@task
def ingest_metadata(refined_metadata):
    pid = random.randint(1,1000)
    logger = get_run_logger()
    response = requests.post(
        url='http://dataverse-importer:8090/importer/',
        json={
            #'doi': f'doi:10.5072/FK2/1{pid}',
            'metadata': refined_metadata,
            'dataverse_information': {
                'base_url': 'https://dataverse.museit.eu',
                'dt_alias': 'transformations',
                'api_token': 'fc66fe9a-c1ec-46c0-a55f-8b2d2636853b'
            }
        }
    )
    logger.info(response.json())
    return response


@task
def transform_image(image, filename, transformation_type, color_format):
    new_filename = f'filename_{str(transformation_type)}'
    files = {'file': (filename, image)}
    response = requests.post(
        url='http://imageconverter:8855/convert',
        params={
            'output_format': transformation_type,
            'color_format': color_format
        },
        files=files,
    )
    return new_filename, response.content


@task
def add_file(ch_file, doi, filename):
    files = {'file': (filename, ch_file)}
    data = {
        'json_data': json.dumps({
            'doi': doi,
            'dataverse_information': {
                'base_url': 'https://dataverse.museit.eu',
                'dt_alias': 'transformations',
                'api_token': 'fc66fe9a-c1ec-46c0-a55f-8b2d2636853b'
            }
        })
    }
    response = requests.post(
        url='http://dataverse-importer:8090/file-upload/',
        files=files,
        data=data
    )
    print(repr(response.request))
    return response

@flow
def ingest_to_dataverse():
    logger = get_run_logger()
    with open('museitmapping.json', 'r') as mapping:
        mappingjson = json.load(mapping)
    with open('museittemplate.json', 'r') as template:
        templatejson = json.load(template)
    with open('test_ingest.xlsx', 'rb') as excel_file:
        workbook = load_workbook(excel_file)
    ws = workbook.active
    pid = 1
    for row in ws.iter_rows(values_only=True):
        metadata = map_line_to_metadata(row=row, mapping=mappingjson, template=templatejson)
        ch_file = retrieve_file_for_metadata(row)
        ingest = ingest_metadata(refined_metadata=metadata.json())
        for transformation_type in ['colormerge', 'colormerge_forced', 'combined_forced']:
            for color_type in ['default', 'fifths']:
                t_filename, t_image = transform_image(
                    image=ch_file,
                    filename=row[3],
                    transformation_type=transformation_type,
                    color_format=color_type
                )
                add_file(ch_file=t_image, doi=ingest.json()['data']['persistentId'], filename=t_filename)
if __name__ == '__main__':
    ingest_to_dataverse()
